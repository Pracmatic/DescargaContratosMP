#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automatiza la descarga y normalización de reportes de contratos.
Requisitos:
  pip install requests pandas openpyxl xlrd

Qué hace:
1) Lee "CONSOLIDADO.xlsx", hoja "Hoja1", tabla "Tabla3" y extrae columnas:
   - "Link" (URL de descarga; si es hipervínculo, se lee el target)
   - "Nombre archivo" (sin extensión; el script añade .xlsx)
2) Descarga cada archivo a la carpeta ./SLEP (la crea si no existe).
3) Convierte/guarda cada archivo como .xlsx con el nombre indicado.
4) Abre cada .xlsx resultante y:
   - Elimina las primeras 10 filas.
   - Asegura que la hoja activa se llame "Hoja1" (y que sea la única).
   - Crea una tabla Excel desde A hasta Y y todas las filas con datos, llamada "Tabla1".
"""

import io
import sys
import re
import time
import shutil
from html.parser import HTMLParser
from pathlib import Path
from typing import List, Tuple, Optional, Iterable

import requests
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# -------------------- Config --------------------
BASE_DIR = Path(__file__).resolve().parent
CONSOLIDADO_PATH = BASE_DIR / "CONSOLIDADO.xlsx"
SALIDA_DIR = BASE_DIR / "SLEP"
HOJA_CONSOLIDADO = "Hoja1"
NOMBRE_TABLA_CONSOLIDADO = "Tabla3"
COL_LINK = "Link"
COL_NOMBRE = "Nombre archivo"
NOMBRE_HOJA_FINAL = "Hoja1"
NOMBRE_TABLA_FINAL = "Tabla1"
COLUMNA_FINAL_TABLA = 25  # A(1) ... Y(25)
HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; autodownload/1.0)"}
TIMEOUT = 60

EXPECTED_COLUMN_NAMES = [
    "Número del contrato",
    "Nombre del contrato",
    "ID licitación / OC",
    "RUT organismo",
    "Nombre organismo",
    "Ejecución del contrato",
    "Categoría del contrato",
    "Tipo de contrato",
    "Unidad requirente",
    "Unidad de moneda",
    "del contrato",
    "ejecutado",
    "por ejecutar",
    "Fecha de inicio",
    "Fecha de término",
    "Estado del contrato",
    "Hitos de pago incumplidos",
    "por vencer",
    "vencidas",
    "cobradas",
    "solicitadas",
    "aplicadas",
    "Días de vigencia",
    "Días restantes",
    "Evaluación",
]

# -------------------- Utilidades --------------------
def safe_filename(name: str) -> str:
    name = name.strip()
    # reemplazar caracteres no válidos en Windows/macOS
    name = re.sub(r'[\\/*?:"<>|]', "_", name)
    # quitar espacios al final
    return name

def read_table_from_excel(path: Path, sheet_name: str, table_name: str) -> Tuple[List[str], List[List[object]]]:
    """
    Lee una tabla estructurada de Excel por nombre usando openpyxl.
    Retorna (headers, rows).
    """
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"No existe la hoja '{sheet_name}' en {path.name}")
        ws = wb[sheet_name]

        if table_name not in ws.tables:
            raise ValueError(f"No existe la tabla '{table_name}' en la hoja '{sheet_name}' de {path.name}")

        table = ws.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)

        # primera fila de la tabla = encabezados
        headers = [ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]

        rows = []
        for r in range(min_row + 1, max_row + 1):
            rows.append([ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)])

        return headers, rows
    finally:
        wb.close()

def extract_links_and_names(headers: List[str], rows: List[List[object]], excel_path: Path) -> List[Tuple[str, str]]:
    """
    A partir de headers/rows de la tabla, extrae (url, nombre_archivo_sin_ext).
    Si la celda de 'Link' es un hipervínculo, volver a abrir con openpyxl para conseguir el target.
    """
    try:
        link_idx = headers.index(COL_LINK)
        name_idx = headers.index(COL_NOMBRE)
    except ValueError as e:
        raise ValueError(f"No se encuentran las columnas requeridas '{COL_LINK}' y '{COL_NOMBRE}' en la tabla.") from e

    # Abrimos el libro para poder leer los hipervínculos reales si existen
    wb = load_workbook(excel_path, data_only=True, read_only=False)
    try:
        ws = wb[HOJA_CONSOLIDADO]
        # Localizar nuevamente los límites de la tabla para conocer fila inicial
        table = ws.tables[NOMBRE_TABLA_CONSOLIDADO]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)

        pares: List[Tuple[str, str]] = []
        for i, row in enumerate(rows, start=0):
            cell_link = ws.cell(row=min_row + 1 + i, column=min_col + link_idx)
            # Si hay hipervínculo, preferimos el target; si no, tomamos el valor literal
            url = None
            if cell_link.hyperlink is not None and getattr(cell_link.hyperlink, "target", None):
                url = cell_link.hyperlink.target
            else:
                # caemos al valor textual
                url = str(row[link_idx]).strip() if row[link_idx] is not None else ""

            name_val = row[name_idx]
            if name_val is None:
                continue
            nombre_archivo = safe_filename(str(name_val))

            if url:
                pares.append((url, nombre_archivo))
        return pares
    finally:
        wb.close()

def download_file(url: str) -> bytes:
    resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT, stream=True)
    resp.raise_for_status()
    return resp.content

def _normalize_whitespace(value: Optional[str]) -> str:
    if value is None:
        return ""
    value = re.sub(r"\s+", " ", value)
    return value.strip()


class _HTMLTableParser(HTMLParser):
    """Extrae tablas simples de un HTML."""

    def __init__(self) -> None:
        super().__init__()
        self.tables: List[dict] = []
        self._table_stack: List[dict] = []
        self._current_row: Optional[dict] = None
        self._current_cell: Optional[dict] = None

    def handle_starttag(self, tag: str, attrs: List[Tuple[str, Optional[str]]]) -> None:
        attrs_dict = {k: v for k, v in attrs}
        if tag == "table":
            table = {"attrs": attrs_dict, "rows": []}
            self.tables.append(table)
            self._table_stack.append(table)
            self._current_row = None
            self._current_cell = None
        elif tag == "tr" and self._table_stack:
            row = {"cells": []}
            self._table_stack[-1]["rows"].append(row)
            self._current_row = row
        elif tag in ("td", "th") and self._current_row is not None:
            cell = {
                "header": tag == "th",
                "text": "",
                "colspan": attrs_dict.get("colspan", "1"),
            }
            self._current_row["cells"].append(cell)
            self._current_cell = cell
        elif tag == "br" and self._current_cell is not None:
            self._current_cell["text"] += "\n"

    def handle_endtag(self, tag: str) -> None:
        if tag == "table" and self._table_stack:
            self._table_stack.pop()
            self._current_row = None
            self._current_cell = None
        elif tag == "tr":
            self._current_row = None
            self._current_cell = None
        elif tag in ("td", "th"):
            self._current_cell = None

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell["text"] += data


def _expand_table_rows(table: dict) -> List[List[str]]:
    rows: List[List[str]] = []
    for raw_row in table.get("rows", []):
        cells = raw_row.get("cells", [])
        if not cells:
            continue
        expanded: List[str] = []
        for cell in cells:
            text = _normalize_whitespace(cell.get("text", ""))
            try:
                span = int(cell.get("colspan", "1") or "1")
            except ValueError:
                span = 1
            span = max(span, 1)
            expanded.extend([text] * span)
        rows.append(expanded)
    return rows


def _select_html_table(tables: Iterable[dict]) -> Optional[List[List[str]]]:
    best: Tuple[int, int, int, int, List[List[str]]] | None = None
    for table in tables:
        rows = _expand_table_rows(table)
        if not rows:
            continue
        text_join = " ".join(cell.lower() for row in rows for cell in row if cell)
        contains_keyword = sum(
            1 for header in ("número del contrato", "nombre del contrato") if header in text_join
        )
        max_cols = max(len(row) for row in rows)
        non_empty_rows = sum(1 for row in rows if any(cell for cell in row))
        table_id = table.get("attrs", {}).get("id", "")
        score = (
            2 if table_id and table_id.lower() == "info" else 0,
            contains_keyword,
            max_cols,
            non_empty_rows,
        )
        if best is None or score > best[:4]:
            best = (*score, rows)  # type: ignore[misc]
    return None if best is None else best[-1]


def _html_bytes_to_dataframe(data: bytes) -> Optional[pd.DataFrame]:
    """
    Intenta interpretar el archivo como HTML (por ejemplo, descargas .xls que en realidad son tablas HTML).
    """
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("latin-1", errors="ignore")

    if "<table" not in text.lower():
        return None

    parser = _HTMLTableParser()
    parser.feed(text)
    parser.close()

    rows = _select_html_table(parser.tables)
    if not rows:
        return None

    # Buscar la fila de encabezados dentro de la tabla seleccionada
    header_index = None
    expected_lower = [name.lower() for name in EXPECTED_COLUMN_NAMES]
    for idx, row in enumerate(rows):
        normalized = [_normalize_whitespace(cell).lower() for cell in row]
        if not normalized:
            continue
        matches = sum(1 for name in expected_lower if name in normalized)
        if matches >= 5 or (
            expected_lower[0] in normalized and expected_lower[1] in normalized
        ):
            header_index = idx
            break

    if header_index is None:
        return None

    data_rows: List[List[Optional[str]]] = []
    for row in rows[header_index + 1 :]:
        cleaned = [_normalize_whitespace(cell) for cell in row]
        if not cleaned or all(cell == "" for cell in cleaned):
            continue
        if cleaned[0].lower() == EXPECTED_COLUMN_NAMES[0].lower():
            # Header repetido dentro de la tabla
            continue
        if any(cell.startswith("</") for cell in cleaned):
            continue

        if len(cleaned) < len(EXPECTED_COLUMN_NAMES):
            cleaned.extend([""] * (len(EXPECTED_COLUMN_NAMES) - len(cleaned)))
        elif len(cleaned) > len(EXPECTED_COLUMN_NAMES):
            cleaned = cleaned[: len(EXPECTED_COLUMN_NAMES)]

        data_rows.append([cell if cell != "" else None for cell in cleaned])

    if not data_rows:
        return None

    df = pd.DataFrame(data_rows, columns=EXPECTED_COLUMN_NAMES)
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def bytes_to_dataframe(data: bytes, source_name: str) -> pd.DataFrame:
    """
    Intenta interpretar los bytes como xlsx, xls o csv/tsv.
    Siempre elimina las primeras 10 filas al cargar (skiprows=10).
    """
    # Primero: intentar .xlsx con openpyxl vía pandas
    try:
        bio = io.BytesIO(data)
        df = pd.read_excel(bio, engine="openpyxl", header=0, skiprows=10)
        return df
    except Exception:
        pass

    # Segundo: intentar .xls con xlrd (requiere xlrd instalado)
    try:
        bio = io.BytesIO(data)
        df = pd.read_excel(bio, engine="xlrd", header=0, skiprows=10)
        return df
    except Exception:
        pass

    # Tercero: intentar HTML con tablas
    try:
        html_df = _html_bytes_to_dataframe(data)
    except Exception:
        html_df = None
    if html_df is not None:
        return html_df

    # Cuarto: intentar CSV/TSV
    for sep in [",", ";", "\t", "|"]:
        try:
            bio = io.BytesIO(data)
            df = pd.read_csv(bio, sep=sep, header=0, skiprows=10, encoding="utf-8", on_bad_lines="skip")
            return df
        except Exception:
            continue

    # Quinto: fallback: intentar sin header y luego asignar encabezados genéricos
    try:
        bio = io.BytesIO(data)
        df = pd.read_csv(bio, header=None, skiprows=10, encoding="utf-8", on_bad_lines="skip")
        # crear encabezados genéricos
        df.columns = [f"Col_{i+1}" for i in range(df.shape[1])]
        return df
    except Exception as e:
        raise ValueError(f"No fue posible interpretar el archivo descargado '{source_name}'.") from e

def ensure_single_sheet_and_name(wb: Workbook, desired_name: str) -> None:
    # asegurar única hoja
    ws = wb.active
    for sh in list(wb.worksheets):
        if sh != ws:
            wb.remove(sh)
    # renombrar
    ws.title = desired_name

def compute_last_data_row(ws, max_col: int) -> int:
    """
    Determina la última fila con datos (no vacía) en el rango de columnas 1..max_col (A..Y).
    """
    last = 1
    for r in range(ws.max_row, 0, -1):
        for c in range(1, max_col + 1):
            if ws.cell(row=r, column=c).value not in (None, "", " "):
                return r
    return last

def add_table_A_to_Y(ws, table_name: str) -> None:
    # Eliminar tabla existente con el mismo nombre si está
    if table_name in ws.tables:
        del ws.tables[table_name]

    # Asegurar que tenemos al menos encabezados en la fila 1 para A..Y
    for c in range(1, 25 + 1):
        cell = ws.cell(row=1, column=c)
        if cell.value in (None, ""):
            # Excel exige encabezados únicos en tablas; si quedan vacíos, Excel asigna "ColumnX" automáticamente.
            # Aquí dejamos vacío y permitimos que Excel lo autogenere visualmente.
            pass

    last_row = compute_last_data_row(ws, 25)
    if last_row < 1:
        last_row = 1

    ref = f"A1:{get_column_letter(25)}{last_row}"
    tbl = Table(displayName=table_name, ref=ref)

    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

def save_dataframe_to_xlsx(df: pd.DataFrame, out_path: Path, sheet_name: str) -> None:
    # Escribimos el dataframe empezando en A1 con encabezados
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    # Reabrimos para aplicar tabla y asegurar hoja única/nombre
    wb = load_workbook(out_path)
    try:
        ensure_single_sheet_and_name(wb, sheet_name)
        ws = wb.active

        # Crear tabla A:Y sobre todas las filas con datos
        add_table_A_to_Y(ws, NOMBRE_TABLA_FINAL)
        wb.save(out_path)
    finally:
        wb.close()

def process_downloaded_bytes(data: bytes, out_xlsx: Path) -> None:
    """
    Convierte bytes a DataFrame (saltando 10 filas) y guarda a .xlsx con tabla.
    """
    df = bytes_to_dataframe(data, out_xlsx.name)

    # Normalizar columnas vacías para que no se pierdan
    # (No agregamos columnas hasta Y aquí; la tabla cubrirá A:Y incluso si hay columnas vacías)
    save_dataframe_to_xlsx(df, out_xlsx, NOMBRE_HOJA_FINAL)

def main():
    print("== Automatización de descarga y formato de contratos ==")
    if not CONSOLIDADO_PATH.exists():
        print(f"ERROR: No se encuentra {CONSOLIDADO_PATH}")
        sys.exit(1)

    SALIDA_DIR.mkdir(exist_ok=True)

    # 1) Leer tabla "Tabla3" de "CONSOLIDADO.xlsx"
    print("Leyendo tabla del CONSOLIDADO...")
    headers, rows = read_table_from_excel(CONSOLIDADO_PATH, HOJA_CONSOLIDADO, NOMBRE_TABLA_CONSOLIDADO)
    pares = extract_links_and_names(headers, rows, CONSOLIDADO_PATH)

    if not pares:
        print("No se encontraron enlaces para procesar.")
        sys.exit(0)

    print(f"Se encontraron {len(pares)} archivos para descargar.\n")

    for idx, (url, nombre_archivo) in enumerate(pares, start=1):
        print(f"[{idx}/{len(pares)}] Descargando: {url}")
        try:
            data = download_file(url)
        except Exception as e:
            print(f"  -> ERROR al descargar: {e}")
            continue

        out_path = SALIDA_DIR / f"{nombre_archivo}.xlsx"

        try:
            process_downloaded_bytes(data, out_path)
            print(f"  -> Guardado y formateado: {out_path.name}")
        except Exception as e:
            print(f"  -> ERROR al procesar/guardar: {e}")
            continue

    print("\nProceso completado. Archivos en:", SALIDA_DIR.resolve())

if __name__ == "__main__":
    main()
